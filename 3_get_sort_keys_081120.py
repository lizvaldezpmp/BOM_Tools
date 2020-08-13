# -*- coding: utf-8 -*-
"""
Created on Sun Jun 28 15:10:28 2020

@author: Liz
"""

import pandas as pd
import numpy as np
#import csv
import sys
import os
#from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import win32com.client



###########################################################################
##### Main Program Starts Here

## This program opens the excel file that has the normalized BOMs with sort keys.
## The input files has the normalized BOMs on separate sheets
## The program adds an additional sheet that has the list of the all unique sort keys
## Run this program after running "normalize_all_bom_types"
        
def get_sort_keys (input_file_name):
    
    print("Processing", input_file_name)
    


    all_sort_keys_sheet_name = "All Sort Keys"
    wb = openpyxl.load_workbook(input_file_name)
    all_sheets = wb.sheetnames
#    print ("sheet names =", all_sheets)
    
    if all_sort_keys_sheet_name not in all_sheets:
        all_sort_keys_sheet = wb.create_sheet(all_sort_keys_sheet_name)
    else:
        all_sort_keys_sheet = wb.remove(wb[all_sort_keys_sheet_name])
        all_sort_keys_sheet = wb.create_sheet(all_sort_keys_sheet_name)

    all_sort_key_list=[]
    all_sort_key_column = 1
    all_sort_key_col_name = "All Sort Keys"
    all_sort_keys_sheet.cell(row=1,column=all_sort_key_column, value=all_sort_keys_sheet_name) 

##############################################################################
### XP Sort Keys
##############################################################################    
    xp_sort_key_list=[]
    xp_sort_column = 2
    xp_sort_key_col_name = "XP Sort Keys"
    

    for sheet in wb:
#        print ("Sheet name is", sheet.title)
        if sheet.title != "All Sort Keys":
            
            num_rows = sheet.max_row-1
#            print("num rows (without 1 header row ) = ", num_rows)

## Start with the second row since the 1st row has the column header
## Remember that the range goes up to but not including the end value (2nd paramter) plus we are starting on row 2

            for which_row in range(2, num_rows+1+1):
                c = sheet.cell(row = which_row, column = xp_sort_column)
#                print("which_row is ", which_row)
#                print("c value = ", c.value)
                cell_as_string = str(c.value)
#                print("cell as string = ", cell_as_string)
                
## Need to convert value to string in order to sort the list
                
                if cell_as_string != "None":
                    if not cell_as_string in xp_sort_key_list:
                        xp_sort_key_list.append(cell_as_string)
                        if not cell_as_string in all_sort_key_list:
                            all_sort_key_list.append(cell_as_string)
#            print("number of items in sort key list = ",len(xp_sort_key_list))
#            print("sort key list is", xp_sort_key_list)
## Sort the list
                
    xp_sort_key_list.sort() 

    list_len = len(xp_sort_key_list)
#    print("length of xp sort_key_list= ", list_len)
    all_sort_keys_sheet.cell(row=1,column=xp_sort_column, value=xp_sort_key_col_name )
## remember that range goes up to but not including the 2nd parameter
    for row_val in range (1, list_len):
#        print("row = ", row_val, "  row value = ", sort_key_list[row_val])
        all_sort_keys_sheet.cell (row=row_val+1, column=xp_sort_column, value=xp_sort_key_list[row_val])

##############################################################################
### Mfg Sort Keys
##############################################################################
        
    
    mfg_sort_key_list=[]
    mfg_sort_column = 3
    mfg_sort_key_col_name = "Mfg Sort Keys"
    

    for sheet in wb:
#        print ("Sheet name is", sheet.title)
        if sheet.title != "All Sort Keys":
            
            num_rows = sheet.max_row-1
#            print("num rows (without 1 header row ) = ", num_rows)

## Start with the second row since the 1st row has the column header
## Remember that the range goes up to but not including the end value (2nd paramter) plus we are starting on row 2

            for which_row in range(2, num_rows+1+1):
                c = sheet.cell(row = which_row, column = mfg_sort_column)
#                print("which_row is ", which_row)
#                print("c value = ", c.value)
                cell_as_string = str(c.value)
#                print("cell as string = ", cell_as_string)
                
## Need to convert value to string in order to sort the list
                
                if cell_as_string != "None":
                    if not cell_as_string in mfg_sort_key_list:
                        mfg_sort_key_list.append(cell_as_string)
                        if not cell_as_string in all_sort_key_list:
                            all_sort_key_list.append(cell_as_string)
#            print("number of items in sort key list = ",len(mfg_sort_key_list))
#            print("sort key list is", mfg_sort_key_list)
## Sort the list
                
    mfg_sort_key_list.sort() 

    list_len = len(mfg_sort_key_list)
#    print("length of mfg sort_key_list= ", list_len)
    all_sort_keys_sheet.cell(row=1,column=mfg_sort_column, value=mfg_sort_key_col_name)
## remember that range goes up to but not including the 2nd parameter
    for row_val in range (1, list_len):
#        print("row = ", row_val, "  row value = ", sort_key_list[row_val])
        all_sort_keys_sheet.cell (row=row_val+1, column=mfg_sort_column, value=mfg_sort_key_list[row_val])
 
##############################################################################
###  Make a column with both Xp and Mfg P/Ns
##############################################################################
    all_sort_key_list.sort()
    list_len = len(all_sort_key_list)
    all_sort_keys_sheet.cell(row=1,column=all_sort_key_column, value=all_sort_key_col_name)
    for row_val in range (1, list_len):
#        print("row = ", row_val, "  row value = ", sort_key_list[row_val])
        all_sort_keys_sheet.cell (row=row_val+1, column=all_sort_key_column, value=all_sort_key_list[row_val])


    print("")
    print("NEXT STEP:  If no errors were reported, use the output file to put in to the Vlookup excel tool")
    
    wb.save(input_file_name)
    


# The following makes this program start running at function above
# when executed as a stand-alone program.    
if __name__ == '__main__':
    
    if len(sys.argv) != 2:
        print ("ERROR:  Re-run program with normalized BOM file")
        sys.exit(1)
    else:
        get_sort_keys (sys.argv[1])
   
    